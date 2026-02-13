import openpyxl
import re
import shutil
from datetime import datetime

SRC = r"c:\Users\Administrator\OneDrive\000_Ai Project\PHP_SQL\ตรวจสอบคุณสมบัติ\ตารางแปลงสอบทาน2.xlsx"
BACKUP = SRC.replace(".xlsx", "_backup.xlsx")
REPORT = r"c:\Users\Administrator\OneDrive\000_Ai Project\PHP_SQL\tools\fix_report.txt"

# Backup first
shutil.copy2(SRC, BACKUP)
print(f"Backup saved: {BACKUP}")

wb = openpyxl.load_workbook(SRC)
ws = wb[wb.sheetnames[0]]  # first sheet

fixes = []

# --- Fix 1: Surname _x000D_ artifact in col 23 (SURNAME) ---
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=23).value
    if val and ('_x000D_' in str(val) or '\r' in str(val) or '\n' in str(val)):
        old = str(val)
        # Clean: remove _x000D_, \r, \n and any trailing garbage
        new = re.sub(r'_x000D_', '', old)
        new = new.replace('\r', '').replace('\n', '').strip()
        ws.cell(row=row, column=23).value = new
        fixes.append(f"[Fix SURNAME] Row {row}: '{old.strip()}' -> '{new}'")

# --- Fix 2: NAME col 22 - same check ---
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=22).value
    if val and ('_x000D_' in str(val) or '\r' in str(val) or '\n' in str(val)):
        old = str(val)
        new = re.sub(r'_x000D_', '', old)
        new = new.replace('\r', '').replace('\n', '').strip()
        ws.cell(row=row, column=22).value = new
        fixes.append(f"[Fix NAME] Row {row}: '{old.strip()}' -> '{new}'")

# --- Fix 3: HOME_NO (col 25) date -> text ---
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=25).value
    if val is None:
        continue
    
    # Check if it's a datetime object (Excel stored as date)
    if isinstance(val, datetime):
        # Excel may have converted house number to date
        # e.g., "30" -> 1900-01-30, "19" -> 2026-01-19
        # The day part is usually the original number
        day = val.day
        month = val.month
        year = val.year
        
        # If year is 1900 or 1899, it's likely a small number that Excel auto-converted
        # Day of 1900-01-XX means the original was just the number XX
        if year == 1900 or year == 1899:
            new_val = str(day)
        else:
            # For other dates, could be like 2026-01-19 -> original was "19" or a real date issue
            # Since HOME_NO should be a house number (text), let's try day
            new_val = str(day)
        
        old_str = val.strftime('%Y-%m-%d')
        ws.cell(row=row, column=25).value = new_val
        fixes.append(f"[Fix HOME_NO] Row {row}: date '{old_str}' -> '{new_val}'")
    
    # Also check string dates like "2026-01-19 00:00:00"
    elif isinstance(val, str) and re.match(r'\d{4}-\d{2}-\d{2}', val):
        match = re.match(r'\d{4}-(\d{2})-(\d{2})', val)
        day = int(match.group(2))
        new_val = str(day)
        ws.cell(row=row, column=25).value = new_val
        fixes.append(f"[Fix HOME_NO] Row {row}: date str '{val}' -> '{new_val}'")

# Save
wb.save(SRC)
wb.close()

# Write report
with open(REPORT, 'w', encoding='utf-8') as f:
    f.write(f"=== Fix Report for ตารางแปลงสอบทาน2.xlsx ===\n")
    f.write(f"Total fixes: {len(fixes)}\n\n")
    for fix in fixes:
        f.write(fix + "\n")

print(f"\nTotal fixes applied: {len(fixes)}")
print(f"Report saved: {REPORT}")
for fix in fixes:
    print(f"  {fix}")
print("\nDone!")
