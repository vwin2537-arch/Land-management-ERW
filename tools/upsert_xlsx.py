"""
UPSERT ข้อมูลจาก ตารางแปลงสอบทาน2.xlsx เข้า DB
- villagers: UPSERT by id_card_number
- land_plots: UPSERT by plot_code (SPAR_CODE)
"""
import openpyxl
import pymysql
import math
import os
import re
import sys
import io

# ============================================================
# Config
# ============================================================
XLSX_PATH = r"c:\Users\Administrator\OneDrive\000_Ai Project\PHP_SQL\ตรวจสอบคุณสมบัติ\ตารางแปลงสอบทาน2.xlsx"
ENV_PATH  = r"c:\Users\Administrator\OneDrive\000_Ai Project\PHP_SQL\.env"
LOG_PATH  = r"c:\Users\Administrator\OneDrive\000_Ai Project\PHP_SQL\tools\upsert_report.txt"

# Open log file + keep stderr for terminal progress
log_file = io.open(LOG_PATH, "w", encoding="utf-8")
original_stderr = sys.stderr

def log(msg):
    log_file.write(msg + "\n")
    log_file.flush()

def progress(msg):
    """Show in both log file and terminal"""
    log(msg)
    try:
        original_stderr.write(msg + "\n")
        original_stderr.flush()
    except:
        pass

# ============================================================
# Read .env
# ============================================================
def read_env(path):
    env = {}
    if not os.path.exists(path):
        return env
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                k, v = line.split('=', 1)
                env[k.strip()] = v.strip()
    return env

env = read_env(ENV_PATH)

# Parse MYSQL_URL if present (same as PHP database.php)
mysql_url = env.get('MYSQL_URL', '') or env.get('MYSQLDATABASE_URL', '')
if mysql_url:
    from urllib.parse import urlparse
    p = urlparse(mysql_url)
    DB_HOST = p.hostname or '127.0.0.1'
    DB_PORT = p.port or 3306
    DB_USER = p.username or 'root'
    DB_PASS = p.password or ''
    DB_NAME = (p.path or '/land_management').lstrip('/')
else:
    DB_HOST = env.get('DB_HOST', '127.0.0.1')
    DB_PORT = int(env.get('DB_PORT', '3306'))
    DB_USER = env.get('DB_USER', 'root')
    DB_PASS = env.get('DB_PASS', '')
    DB_NAME = env.get('DB_NAME', 'land_management')

# ============================================================
# UTM Zone 47N -> WGS84 Lat/Lng
# ============================================================
def utm_to_latlng(easting, northing, zone=47, northern=True):
    a = 6378137.0
    f = 1 / 298.257223563
    e = math.sqrt(2 * f - f * f)
    e2 = e * e / (1 - e * e)
    k0 = 0.9996
    x = easting - 500000.0
    y = northing if northern else northing - 10000000.0

    M = y / k0
    mu = M / (a * (1 - e*e/4 - 3*e**4/64 - 5*e**6/256))
    e1 = (1 - math.sqrt(1 - e*e)) / (1 + math.sqrt(1 - e*e))

    phi1 = (mu
        + (3*e1/2 - 27*e1**3/32) * math.sin(2*mu)
        + (21*e1**2/16 - 55*e1**4/32) * math.sin(4*mu)
        + (151*e1**3/96) * math.sin(6*mu)
        + (1097*e1**4/512) * math.sin(8*mu))

    C1 = e2 * math.cos(phi1)**2
    T1 = math.tan(phi1)**2
    N1 = a / math.sqrt(1 - e*e * math.sin(phi1)**2)
    R1 = a * (1 - e*e) / (1 - e*e * math.sin(phi1)**2)**1.5
    D = x / (N1 * k0)

    lat = phi1 - (N1 * math.tan(phi1) / R1) * (
        D**2/2
        - (5 + 3*T1 + 10*C1 - 4*C1**2 - 9*e2) * D**4/24
        + (61 + 90*T1 + 298*C1 + 45*T1**2 - 252*e2 - 3*C1**2) * D**6/720
    )

    lng0 = math.radians((zone - 1) * 6 - 180 + 3)
    lng = lng0 + (D - (1 + 2*T1 + C1) * D**3/6
        + (5 - 2*C1 + 28*T1 - 3*C1**2 + 8*e2 + 24*T1**2) * D**5/120) / math.cos(phi1)

    return round(math.degrees(lat), 7), round(math.degrees(lng), 7)

# ============================================================
# Map PTYPE -> land_use_type enum
# ============================================================
def map_land_use(ptype):
    if not ptype:
        return 'other'
    if 'อยู่อาศัย' in ptype and 'ทำกิน' in ptype:
        return 'mixed'
    if 'อยู่อาศัย' in ptype:
        return 'residential'
    if 'เกษตร' in ptype or 'ทำกิน' in ptype:
        return 'agriculture'
    if 'สวน' in ptype:
        return 'garden'
    if 'ปศุสัตว์' in ptype or 'เลี้ยง' in ptype:
        return 'livestock'
    return 'other'

# ============================================================
# Map REMARK -> remark_risk enum
# ============================================================
def map_remark(remark):
    if not remark:
        return 'not_risky'
    r = str(remark).strip()
    if r == 'ล่อแหลมมีคดี':
        return 'risky_case'
    if r == 'ไม่ล่อแหลมมีคดี':
        return 'not_risky_case'
    if 'ล่อแหลม' in r:
        return 'risky'
    if 'ไม่ล่อแหลม' in r:
        return 'not_risky'
    return 'not_risky'

# ============================================================
# Validate Thai ID
# ============================================================
def check_idcard(idc):
    if not idc or len(str(idc)) != 13 or not str(idc).isdigit():
        return False
    s = sum(int(str(idc)[i]) * (13 - i) for i in range(12))
    return (11 - (s % 11)) % 10 == int(str(idc)[12])

# ============================================================
# Read XLSX headers
# ============================================================
progress("=== UPSERT ตารางแปลงสอบทาน2.xlsx → DB ===")
progress(f"DB: {DB_HOST}:{DB_PORT}/{DB_NAME}")

progress("Loading xlsx...")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb[wb.sheetnames[0]]
progress(f"Sheet: {wb.sheetnames[0]}, rows={ws.max_row}, cols={ws.max_column}")

# Preload all data into list of lists for fast access
all_rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    all_rows.append(list(row))
progress(f"Loaded {len(all_rows)} rows into memory")
wb.close()

# Find header row (scan first 3 rows)
headers = {}
header_row_idx = 0  # 0-based index
for ri in range(min(3, len(all_rows))):
    for ci, val in enumerate(all_rows[ri]):
        if val and str(val).strip().upper() in ('NAME', 'SURNAME', 'IDCARD', 'SPAR_CODE', 'NUM_APAR'):
            header_row_idx = ri
            break
    if header_row_idx != 0:
        break

# Map column names to 0-based indices
for ci, val in enumerate(all_rows[header_row_idx]):
    if val:
        headers[str(val).strip().upper()] = ci

progress(f"Header row: {header_row_idx + 1}")
progress(f"Columns found: {len(headers)}")
for k, v in headers.items():
    log(f"  {k} = col {v + 1}")

# Helper to get cell value from preloaded data
def get_val(row_data, col_name):
    idx = headers.get(col_name)
    if idx is None or idx >= len(row_data):
        return None
    return row_data[idx]

def get_str(row_data, col_name):
    v = get_val(row_data, col_name)
    if v is None:
        return ''
    return str(v).strip()

def get_float(row_data, col_name):
    v = get_val(row_data, col_name)
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0

# Data rows (skip header)
data_rows = all_rows[header_row_idx + 1:]

# ============================================================
# Connect DB
# ============================================================
progress(f"Connecting to {DB_HOST}:{DB_PORT}/{DB_NAME} ...")
conn = pymysql.connect(
    host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
    database=DB_NAME, charset='utf8mb4', autocommit=False, connect_timeout=10
)
cur = conn.cursor()
progress("Connected!")

# ============================================================
# Process rows
# ============================================================
total_rows = len(data_rows)
progress(f"Data rows: {total_rows}")

stats = {
    'villager_insert': 0, 'villager_update': 0,
    'plot_insert': 0, 'plot_update': 0,
    'errors': 0, 'skipped': 0
}
errors = []

try:
    for row_num, rd in enumerate(data_rows, start=1):

        # --- Read fields ---
        idcard     = get_str(rd, 'IDCARD')
        name_title = get_str(rd, 'NAME_TITLE')
        name       = get_str(rd, 'NAME')
        surname    = get_str(rd, 'SURNAME')
        home_no    = get_str(rd, 'HOME_NO')
        home_ban   = get_str(rd, 'HOME_BAN')
        home_moo   = get_str(rd, 'HOME_MOO')
        home_tam   = get_str(rd, 'HOME_TAM')
        home_amp   = get_str(rd, 'HOME_AMP')
        home_prov  = get_str(rd, 'HOME_PROV')

        spar_code  = get_str(rd, 'SPAR_CODE')
        name_dnp   = get_str(rd, 'NAME_DNP')
        code_dnp   = get_str(rd, 'CODE_DNP')
        apar_code  = get_str(rd, 'APAR_CODE')
        apar_no    = get_str(rd, 'APAR_NO')
        num_apar   = get_str(rd, 'NUM_APAR')
        spar_no    = get_str(rd, 'SPAR_NO').zfill(5) if get_str(rd, 'SPAR_NO') else ''
        num_spar   = get_str(rd, 'NUM_SPAR').zfill(5) if get_str(rd, 'NUM_SPAR') else ''
        par_ban    = get_str(rd, 'PAR_BAN')
        ban_e      = get_str(rd, 'BAN_E')
        par_moo    = get_str(rd, 'PAR_MOO')
        par_tam    = get_str(rd, 'PAR_TAM')
        par_amp    = get_str(rd, 'PAR_AMP')
        par_prov   = get_str(rd, 'PAR_PROV')
        perimeter  = get_float(rd, 'PERIMETER')
        rai        = get_float(rd, 'RAI')
        ngan       = get_float(rd, 'NGAN')
        wa_sq      = get_float(rd, 'WA_SQ')
        area_rai   = get_float(rd, 'AREA_RAI')
        ptype      = get_str(rd, 'PTYPE')
        remark     = get_str(rd, 'REMARK')
        ban_type   = get_str(rd, 'BAN_TYPE')
        year_val   = get_str(rd, 'YEAR')

        e_val      = get_float(rd, 'E')
        n_val      = get_float(rd, 'N')
        target_fid_str = get_str(rd, 'TARGET_FID')

        # Skip empty rows
        if not idcard and not spar_code and not name:
            stats['skipped'] += 1
            continue

        # --- Data issues ---
        issues = []
        if not idcard:
            issues.append('ไม่มีเลขบัตร')
        elif not check_idcard(idcard):
            issues.append(f'เลขบัตรไม่ถูกต้อง: {idcard}')
        if not name:
            issues.append('ไม่มีชื่อ')
        if not surname:
            issues.append('ไม่มีนามสกุล')

        # --- UTM -> LatLng ---
        lat, lng = None, None
        if e_val > 0 and n_val > 0:
            lat, lng = utm_to_latlng(e_val, n_val, 47, True)

        # --- Year conversion (พ.ศ. -> ค.ศ.) ---
        occupation_since = None
        if year_val:
            try:
                yr = int(float(year_val))
                if yr > 2400:
                    yr -= 543
                occupation_since = yr
            except (ValueError, TypeError):
                pass

        # --- Plot code ---
        plot_code = spar_code
        if not plot_code:
            plot_code = f'IMP-{row_num:05d}'
            issues.append(f'ไม่มี SPAR_CODE — ใช้รหัส {plot_code}')

        # --- Land use ---
        land_use_type = map_land_use(ptype)
        remark_risk = map_remark(remark)

        # --- Address ---
        home_addr = f"{home_no} หมู่ {home_moo or '-'}" if home_no else ''

        # --- Status ---
        issue_text = '; '.join(issues) if issues else None
        status = 'pending_review' if issue_text else 'surveyed'

        # --- Target FID ---
        target_fid = None
        if target_fid_str:
            try:
                target_fid = int(float(target_fid_str))
            except (ValueError, TypeError):
                pass

        # --- BAN_TYPE ---
        ban_type_val = None
        if ban_type:
            try:
                ban_type_val = str(int(float(ban_type)))
            except (ValueError, TypeError):
                ban_type_val = ban_type

        # Use RAI if > 0, otherwise AREA_RAI
        final_rai = rai if rai > 0 else area_rai

        # ============================================================
        # UPSERT Villager
        # ============================================================
        villager_id = None
        if idcard and check_idcard(idcard):
            # Check if exists
            cur.execute("SELECT villager_id FROM villagers WHERE id_card_number = %s LIMIT 1", (idcard,))
            existing = cur.fetchone()

            if existing:
                villager_id = existing[0]
                cur.execute("""
                    UPDATE villagers SET
                        prefix = COALESCE(%s, prefix),
                        first_name = COALESCE(%s, first_name),
                        last_name = COALESCE(%s, last_name),
                        village_name = COALESCE(%s, village_name),
                        village_no = COALESCE(%s, village_no),
                        sub_district = COALESCE(%s, sub_district),
                        district = COALESCE(%s, district),
                        province = COALESCE(%s, province),
                        address = COALESCE(%s, address)
                    WHERE id_card_number = %s
                """, (
                    name_title or None, name or None, surname or None,
                    home_ban or None, home_moo or None,
                    home_tam or None, home_amp or None, home_prov or None,
                    home_addr or None, idcard
                ))
                stats['villager_update'] += 1
            else:
                cur.execute("""
                    INSERT INTO villagers
                        (id_card_number, prefix, first_name, last_name,
                         village_name, village_no, sub_district, district, province, address)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    idcard, name_title or None, name or 'ไม่ระบุ', surname or 'ไม่ระบุ',
                    home_ban or None, home_moo or None,
                    home_tam or None, home_amp or None, home_prov or None,
                    home_addr or None
                ))
                villager_id = cur.lastrowid
                stats['villager_insert'] += 1
        elif idcard:
            # Bad ID card - still try to find or create
            placeholder = f'BAD_{idcard}'[:13]
            cur.execute("SELECT villager_id FROM villagers WHERE id_card_number = %s LIMIT 1", (idcard,))
            existing = cur.fetchone()
            if existing:
                villager_id = existing[0]
                stats['villager_update'] += 1
            else:
                cur.execute("""
                    INSERT INTO villagers
                        (id_card_number, prefix, first_name, last_name,
                         village_name, village_no, sub_district, district, province, address)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    idcard, name_title or None, name or 'ไม่ระบุ', surname or 'ไม่ระบุ',
                    home_ban or None, home_moo or None,
                    home_tam or None, home_amp or None, home_prov or None,
                    home_addr or None
                ))
                villager_id = cur.lastrowid
                stats['villager_insert'] += 1
        else:
            # No IDCARD at all
            temp_idc = f'TEMP_{row_num:05d}'
            cur.execute("""
                INSERT INTO villagers
                    (id_card_number, prefix, first_name, last_name,
                     village_name, village_no, sub_district, district, province, address)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                temp_idc, name_title or None, name or f'ไม่ระบุ_{row_num}', surname or f'ไม่ระบุ_{row_num}',
                home_ban or None, home_moo or None,
                home_tam or None, home_amp or None, home_prov or None,
                home_addr or None
            ))
            villager_id = cur.lastrowid
            stats['villager_insert'] += 1

        # ============================================================
        # UPSERT Land Plot
        # ============================================================
        cur.execute("SELECT plot_id FROM land_plots WHERE plot_code = %s LIMIT 1", (plot_code,))
        existing_plot = cur.fetchone()

        if existing_plot:
            # UPDATE existing plot
            cur.execute("""
                UPDATE land_plots SET
                    villager_id = %s,
                    park_name = COALESCE(%s, park_name),
                    area_rai = %s, area_ngan = %s, area_sqwa = %s,
                    land_use_type = %s,
                    latitude = COALESCE(%s, latitude),
                    longitude = COALESCE(%s, longitude),
                    status = %s,
                    code_dnp = COALESCE(%s, code_dnp),
                    apar_code = COALESCE(%s, apar_code),
                    apar_no = COALESCE(%s, apar_no),
                    num_apar = COALESCE(%s, num_apar),
                    spar_code = COALESCE(%s, spar_code),
                    ban_e = COALESCE(%s, ban_e),
                    perimeter = %s,
                    ban_type = COALESCE(%s, ban_type),
                    num_spar = COALESCE(%s, num_spar),
                    spar_no = COALESCE(%s, spar_no),
                    par_ban = COALESCE(%s, par_ban),
                    par_moo = COALESCE(%s, par_moo),
                    par_tam = COALESCE(%s, par_tam),
                    par_amp = COALESCE(%s, par_amp),
                    par_prov = COALESCE(%s, par_prov),
                    ptype = COALESCE(%s, ptype),
                    target_fid = COALESCE(%s, target_fid),
                    occupation_since = COALESCE(%s, occupation_since),
                    remark_risk = %s,
                    data_issues = %s
                WHERE plot_code = %s
            """, (
                villager_id,
                name_dnp or None,
                final_rai, ngan, wa_sq,
                land_use_type,
                lat, lng,
                status,
                code_dnp or None, apar_code or None, apar_no or None,
                num_apar or None, spar_code or None, ban_e or None,
                perimeter, ban_type_val,
                num_spar or None, spar_no or None,
                par_ban or None, par_moo or None, par_tam or None,
                par_amp or None, par_prov or None, ptype or None,
                target_fid, occupation_since,
                remark_risk, issue_text,
                plot_code
            ))
            stats['plot_update'] += 1
        else:
            # INSERT new plot
            cur.execute("""
                INSERT INTO land_plots
                    (plot_code, villager_id, park_name, zone,
                     area_rai, area_ngan, area_sqwa,
                     land_use_type, latitude, longitude, status, notes,
                     code_dnp, apar_code, apar_no, num_apar, spar_code,
                     ban_e, perimeter, ban_type, num_spar, spar_no,
                     par_ban, par_moo, par_tam, par_amp, par_prov,
                     ptype, target_fid, occupation_since, remark_risk, data_issues)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                plot_code, villager_id, name_dnp or None, None,
                final_rai, ngan, wa_sq,
                land_use_type, lat, lng, status, None,
                code_dnp or None, apar_code or None, apar_no or None,
                num_apar or None, spar_code or None,
                ban_e or None, perimeter, ban_type_val,
                num_spar or None, spar_no or None,
                par_ban or None, par_moo or None, par_tam or None,
                par_amp or None, par_prov or None,
                ptype or None, target_fid, occupation_since,
                remark_risk, issue_text
            ))
            stats['plot_insert'] += 1

        if row_num % 200 == 0:
            progress(f"  Processing {row_num}/{total_rows} ...")

    conn.commit()
    progress(f"\n{'='*50}")
    progress(f"✅ UPSERT สำเร็จ!")
    progress(f"   ราษฎรสร้างใหม่:    {stats['villager_insert']}")
    progress(f"   ราษฎรอัพเดท:       {stats['villager_update']}")
    progress(f"   แปลงสร้างใหม่:     {stats['plot_insert']}")
    progress(f"   แปลงอัพเดท:        {stats['plot_update']}")
    progress(f"   ข้ามแถวว่าง:       {stats['skipped']}")
    progress(f"   ข้อผิดพลาด:        {stats['errors']}")

    # Summary from DB
    cur.execute("SELECT COUNT(*) FROM villagers")
    progress(f"\n   [DB] villagers ทั้งหมด: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM land_plots")
    progress(f"   [DB] land_plots ทั้งหมด: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM land_plots WHERE data_issues IS NOT NULL")
    progress(f"   [DB] แปลงมีปัญหา: {cur.fetchone()[0]}")
    progress(f"{'='*50}")

except Exception as ex:
    conn.rollback()
    progress(f"\n❌ Error at row {row_num}: {ex}")
    import traceback
    traceback.print_exc(file=log_file)
finally:
    cur.close()
    conn.close()

progress("\nDone!")
log_file.close()
