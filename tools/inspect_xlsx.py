import openpyxl
import re
import sys
import io

# Redirect stdout to file with UTF-8
OUTPUT_FILE = r"c:\Users\Administrator\OneDrive\000_Ai Project\PHP_SQL\tools\inspect_result.txt"
sys.stdout = io.open(OUTPUT_FILE, "w", encoding="utf-8")

def check_thai_id(id_str):
    """ตรวจสอบเลขบัตรประชาชนไทย 13 หลัก"""
    issues = []
    if not id_str:
        issues.append("ไม่มีเลขบัตร")
        return issues
    
    # Clean up - remove spaces, dashes
    clean = str(id_str).strip().replace(" ", "").replace("-", "").replace("\t", "")
    
    # Check if numeric
    if not clean.isdigit():
        issues.append(f"มีอักขระที่ไม่ใช่ตัวเลข: '{id_str}'")
        return issues
    
    # Check length
    if len(clean) != 13:
        issues.append(f"ไม่ครบ 13 หลัก (มี {len(clean)} หลัก): '{clean}'")
        return issues
    
    # Check digit validation (Thai national ID checksum)
    total = 0
    for i in range(12):
        total += int(clean[i]) * (13 - i)
    remainder = total % 11
    check_digit = (11 - remainder) % 10
    
    if int(clean[12]) != check_digit:
        issues.append(f"Check digit ไม่ถูกต้อง (คาดหวัง {check_digit}, ได้ {clean[12]}): '{clean}'")
    
    # Check starts with 0 or unusual prefix
    if clean[0] == '0':
        issues.append(f"ขึ้นต้นด้วย 0 (ผิดปกติ): '{clean}'")
    
    return issues

def check_name(name_str, col_label):
    """ตรวจสอบชื่อ-สกุล"""
    issues = []
    if not name_str or str(name_str).strip() == "":
        issues.append(f"{col_label}: ว่างเปล่า")
        return issues
    
    name = str(name_str).strip()
    
    # Check for numbers in name
    if re.search(r'\d', name):
        issues.append(f"{col_label}: มีตัวเลขปนในชื่อ '{name}'")
    
    # Check for _x000D_ or carriage return artifacts
    if '_x000D_' in name or '\r' in name or '\n' in name:
        issues.append(f"{col_label}: มี carriage return/newline artifact '{repr(name)}'")
    
    # Check for unusual characters
    if re.search(r'[!@#$%^&*()=+\[\]{}<>|\\/:;]', name):
        issues.append(f"{col_label}: มีอักขระพิเศษ '{name}'")
    
    # Check very short name (less than 2 chars)
    if len(name) < 2:
        issues.append(f"{col_label}: สั้นเกินไป '{name}'")
    
    # Check if name has only spaces or dots
    if re.match(r'^[\s.]+$', name):
        issues.append(f"{col_label}: มีแต่จุดหรือช่องว่าง '{name}'")
    
    return issues

def main():
    filepath = r"c:\Users\Administrator\OneDrive\000_Ai Project\PHP_SQL\ตรวจสอบคุณสมบัติ\ตารางแปลงสอบทาน2.xlsx"
    
    print(f"=== กำลังอ่านไฟล์: ตารางแปลงสอบทาน2.xlsx ===\n")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*60}")
        
        ws = wb[sheet_name]
        
        # Show first few rows to understand structure
        print(f"\nจำนวนแถว: {ws.max_row}, จำนวนคอลัมน์: {ws.max_column}")
        
        # Print header rows (first 3 rows)
        print("\n--- Header (แถว 1-3) ---")
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    row_data.append(f"[{col_idx}]{val}")
            if row_data:
                print(f"  แถว {row_idx}: {' | '.join(row_data)}")
        
        # Try to find ID card column and name columns
        # Scan header rows for keywords
        id_col = None
        fname_col = None
        lname_col = None
        name_col = None  # combined name column
        title_col = None
        
        for row_idx in range(1, min(5, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                val_str = str(val).strip().upper()
                
                if any(k in val_str for k in ['เลขบัตร', 'บัตรประชาชน', 'เลขประจำตัว', 'IDCARD', 'ID_CARD']):
                    id_col = col_idx
                if val_str == 'NAME' or any(k in str(val).strip() for k in ['ชื่อ-สกุล', 'ชื่อ-นามสกุล', 'ชื่อสกุล']):
                    if 'SURNAME' not in val_str and 'สกุล' not in str(val).strip():
                        fname_col = col_idx
                    elif 'SURNAME' not in val_str:
                        name_col = col_idx
                if val_str == 'SURNAME' or any(k in str(val).strip() for k in ['สกุล', 'นามสกุล']):
                    lname_col = col_idx
                if val_str in ('NAME_TITLE', 'TITLE') or 'คำนำหน้า' in str(val).strip():
                    title_col = col_idx
        
        print(f"\n--- คอลัมน์ที่พบ ---")
        print(f"  เลขบัตร: คอลัมน์ {id_col}")
        print(f"  ชื่อ: คอลัมน์ {fname_col}")
        print(f"  สกุล: คอลัมน์ {lname_col}")
        print(f"  ชื่อ-สกุล (รวม): คอลัมน์ {name_col}")
        
        # Now scan data rows
        data_start = 2  # assume header is row 1
        # If header seems multi-row, adjust
        for r in range(1, min(6, ws.max_row + 1)):
            first_val = ws.cell(row=r, column=1).value
            if first_val is not None and (str(first_val).strip().isdigit() or str(first_val).strip() == '1'):
                # Could be data start
                if r > 1:
                    data_start = r
                break
        
        print(f"  Data start row: {data_start}")
        
        # Collect all data and check
        anomalies = []
        all_ids = {}  # track duplicates
        row_count = 0
        
        for row_idx in range(data_start, ws.max_row + 1):
            # Skip empty rows
            row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            if all(v is None for v in row_vals):
                continue
            
            row_count += 1
            row_issues = []
            
            # Check ID
            if id_col:
                id_val = ws.cell(row=row_idx, column=id_col).value
                if id_val is not None:
                    id_str = str(id_val).strip()
                    id_issues = check_thai_id(id_str)
                    row_issues.extend(id_issues)
                    
                    # Track duplicates
                    clean_id = id_str.replace(" ", "").replace("-", "")
                    if clean_id in all_ids:
                        row_issues.append(f"เลขบัตรซ้ำกับแถว {all_ids[clean_id]}")
                    else:
                        all_ids[clean_id] = row_idx
            
            # Check names
            if fname_col:
                fname = ws.cell(row=row_idx, column=fname_col).value
                row_issues.extend(check_name(fname, "ชื่อ"))
            
            if lname_col:
                lname = ws.cell(row=row_idx, column=lname_col).value
                row_issues.extend(check_name(lname, "สกุล"))
            
            if name_col:
                full_name = ws.cell(row=row_idx, column=name_col).value
                row_issues.extend(check_name(full_name, "ชื่อ-สกุล"))
            
            if row_issues:
                # Get display info
                display_parts = [f"แถว {row_idx}"]
                if id_col:
                    display_parts.append(f"เลขบัตร: {ws.cell(row=row_idx, column=id_col).value}")
                if fname_col:
                    display_parts.append(f"ชื่อ: {ws.cell(row=row_idx, column=fname_col).value}")
                if lname_col:
                    display_parts.append(f"สกุล: {ws.cell(row=row_idx, column=lname_col).value}")
                if name_col:
                    display_parts.append(f"ชื่อ-สกุล: {ws.cell(row=row_idx, column=name_col).value}")
                
                anomalies.append({
                    'row': row_idx,
                    'display': ' | '.join(display_parts),
                    'issues': row_issues
                })
        
        print(f"\n  จำนวนแถวข้อมูล: {row_count}")
        
        # Print all data for review (first 20 rows)
        print(f"\n--- ตัวอย่างข้อมูล (แถวแรก 30 แถว) ---")
        shown = 0
        for row_idx in range(data_start, ws.max_row + 1):
            row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            if all(v is None for v in row_vals):
                continue
            shown += 1
            if shown > 30:
                print(f"  ... (เหลืออีก {row_count - 30} แถว)")
                break
            
            parts = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=row_idx, column=c).value
                if v is not None:
                    parts.append(f"[{c}]{v}")
            print(f"  แถว {row_idx}: {' | '.join(parts)}")
        
        # Report anomalies
        print(f"\n{'='*60}")
        print(f"=== ผลการตรวจสอบ Sheet: {sheet_name} ===")
        print(f"{'='*60}")
        
        if anomalies:
            print(f"\n⚠️  พบความผิดปกติ {len(anomalies)} รายการ:\n")
            for a in anomalies:
                print(f"  📌 {a['display']}")
                for issue in a['issues']:
                    print(f"     ❌ {issue}")
                print()
        else:
            print(f"\n✅ ไม่พบความผิดปกติ")
    
    wb.close()

if __name__ == "__main__":
    main()
