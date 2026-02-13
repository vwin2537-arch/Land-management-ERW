"""Read reference form templates (xlsx) to understand official government format"""
import openpyxl
import os, sys

REF_DIR = r"c:\Users\Administrator\OneDrive\000_Ai Project\PHP_SQL\references\แบบฟอร์ม"

out = []
def p(s=""): out.append(str(s))

for fname in os.listdir(REF_DIR):
    fpath = os.path.join(REF_DIR, fname)
    if not fname.endswith('.xlsx'):
        p(f"\n{'='*60}")
        p(f"SKIP (not xlsx): {fname}")
        continue
    
    p(f"\n{'='*60}")
    p(f"FILE: {fname}")
    p(f"{'='*60}")
    
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            p(f"\n--- Sheet: {sname} (rows={ws.max_row}, cols={ws.max_column}) ---")
            
            # Read all rows (up to 50) to understand structure
            for row_idx, row in enumerate(ws.iter_rows(max_row=min(ws.max_row, 50), values_only=False), 1):
                vals = []
                for cell in row:
                    v = cell.value
                    if v is not None:
                        v_str = str(v).strip().replace('\n', ' | ')
                        if len(v_str) > 60:
                            v_str = v_str[:60] + '...'
                        vals.append(f"[{cell.column_letter}]{v_str}")
                if vals:
                    p(f"  R{row_idx}: {' | '.join(vals)}")
            
            # Check merged cells
            if ws.merged_cells.ranges:
                p(f"\n  Merged cells:")
                for mc in ws.merged_cells.ranges:
                    p(f"    {mc}")
        wb.close()
    except Exception as e:
        p(f"  ERROR: {e}")

result = '\n'.join(out)
print(result[:3000])  # Print first part to terminal

# Save full output
outpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ref_forms_structure.txt')
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(result)
print(f"\n\nFull output saved to: {outpath} ({len(result)} chars)")
